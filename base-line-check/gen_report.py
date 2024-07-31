import os

import openpyxl

EXCEL_FILE = "linux_check.xlsx"
SHEET_NAME = "Sheet1"


# 打开Excel文件
wb = openpyxl.load_workbook("linux_check.xlsx")
# 获取指定Sheet
sheet = wb["Sheet1"]
# # 获取Sheet的行数和列数
# max_row = sheet.max_row
# max_col = sheet.max_column
#
# # 遍历Sheet的所有单元格，并输出内容
# for i in range(1, max_row + 1):
#     for j in range(1, max_col + 1):
#         cell_value = sheet.cell(row=i, column=j).value
#         if cell_value:
#             print(f'Cell({i},{j}): {cell_value}')

# 打开Excel文件
# wb = openpyxl.load_workbook('${EXCEL_FILE}')
# 获取指定Sheet
# sheet = wb['${SHEET_NAME}']

# 写入内容到指定单元格
# sheet.cell(row=1, column=1, value='Hello, world!')
# sheet.cell(row=2, column=1, value='This is a test.')

# 保存Excel文件
wb.save("linux_check.xlsx")
